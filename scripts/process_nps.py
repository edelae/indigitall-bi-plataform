"""Process NPS/satisfaction survey CSVs from inDigitall into a clean Excel file.

Reads all CSV files from NPS/ folder, parses the JSON in the 'content' column,
and extracts survey responses into structured columns.

Output: NPS/encuestas_procesadas.xlsx
"""

import csv
import json
import os
from datetime import datetime
from pathlib import Path

# pip install openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

BASE_DIR = Path(__file__).resolve().parent.parent / "NPS"

# Survey question short names
QUESTION_MAP = {
    "¿Cómo calificarías la atención que recibiste?": "Calificacion Atencion",
    "¿La atención fue rápida?": "Atencion Rapida",
    "¿Pudiste resolver tu duda o problema?": "Problema Resuelto",
    "¿Volverías a utilizar este canal para ser atendido?": "Volveria a Usar",
    "¿Cómo calificarías el trato que recibiste por parte del asesor?": "Calificacion Asesor",
    "¿Quieres dejar un comentario o sugerencia? ✍️": "Comentario",
}

# Satisfaction score mapping
SATISFACTION_SCORES = {
    "😊 Muy satisfecho": 5,
    "🙂 Satisfecho": 4,
    "😐 Neutral": 3,
    "😞 Insatisfecho": 2,
    "😡 Muy insatisfecho": 1,
}

YES_NO_SCORES = {
    "⏱️ Sí ": 1,  # note trailing space
    "⏱️ Sí": 1,
    "✅ Sí": 1,
    "👍 Sí": 1,
    "🕓 No": 0,
    "❌ No ": 0,  # note trailing space
    "❌ No": 0,
    "👎 No": 0,
    "Sí": 1,
    "No": 0,
}

MONTH_MAP = {
    "1_JUL": "2025-07",
    "2_AGO": "2025-08",
    "3_SEP": "2025-09",
    "4_OCT": "2025-10",
    "5_NOV": "2025-11",
    "6_DIC": "2025-12",
    "7_ENE": "2026-01",
    "8_FEB": "2026-02",
    "9_MAR": "2026-03",
}


def parse_content(content_str: str) -> dict:
    """Parse the JSON content field and extract survey responses."""
    if not content_str:
        return {}
    try:
        data = json.loads(content_str)
        flow = data.get("eventParameters", {}).get("flowResponse", {})
        return flow
    except (json.JSONDecodeError, TypeError):
        return {}


def extract_survey(flow: dict) -> dict | None:
    """Extract structured survey data from flowResponse dict."""
    if not flow or flow.get("type") != "encuesta":
        return None

    result = {}
    for full_q, short_name in QUESTION_MAP.items():
        val = flow.get(full_q, "")
        result[short_name] = val.strip() if val else ""

    result["Entidad"] = flow.get("entity", "").strip()

    # Compute numeric scores
    result["Score Atencion"] = SATISFACTION_SCORES.get(result["Calificacion Atencion"], "")
    result["Score Asesor"] = SATISFACTION_SCORES.get(result["Calificacion Asesor"], "")
    result["Rapida (1/0)"] = YES_NO_SCORES.get(result["Atencion Rapida"], "")
    result["Resuelto (1/0)"] = YES_NO_SCORES.get(result["Problema Resuelto"], "")
    result["Volveria (1/0)"] = YES_NO_SCORES.get(result["Volveria a Usar"], "")

    # NPS category based on average satisfaction score
    scores = [v for v in [result["Score Atencion"], result["Score Asesor"]] if isinstance(v, (int, float))]
    if scores:
        avg = sum(scores) / len(scores)
        if avg >= 4.5:
            result["NPS Categoria"] = "Promotor"
        elif avg >= 3.0:
            result["NPS Categoria"] = "Pasivo"
        else:
            result["NPS Categoria"] = "Detractor"
    else:
        result["NPS Categoria"] = ""

    return result


def process_all_csvs():
    """Read all CSVs and return list of processed rows."""
    rows = []
    csv_files = sorted(BASE_DIR.glob("*.csv"))

    for csv_file in csv_files:
        stem = csv_file.stem  # e.g. "1_JUL"
        month_label = MONTH_MAP.get(stem, stem)

        with open(csv_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                content = row.get("content", "")
                flow = parse_content(content)
                survey = extract_survey(flow)

                if not survey:
                    continue  # skip non-survey rows

                # Parse date
                msg_date = row.get("messageDate", "")
                try:
                    dt = datetime.fromisoformat(msg_date.replace("+00:00", "+00:00"))
                    date_str = dt.strftime("%Y-%m-%d")
                    time_str = dt.strftime("%H:%M")
                    weekday = ["Lun", "Mar", "Mie", "Jue", "Vie", "Sab", "Dom"][dt.weekday()]
                except Exception:
                    date_str = msg_date[:10] if msg_date else ""
                    time_str = ""
                    weekday = ""

                processed = {
                    "Mes": month_label,
                    "Fecha": date_str,
                    "Hora": time_str,
                    "Dia Semana": weekday,
                    "Contacto": row.get("profileName", ""),
                    "Telefono": row.get("contactId", ""),
                    "Entidad": survey["Entidad"],
                    "Calificacion Atencion": survey["Calificacion Atencion"],
                    "Score Atencion": survey["Score Atencion"],
                    "Calificacion Asesor": survey["Calificacion Asesor"],
                    "Score Asesor": survey["Score Asesor"],
                    "Atencion Rapida": survey["Atencion Rapida"],
                    "Rapida (1/0)": survey["Rapida (1/0)"],
                    "Problema Resuelto": survey["Problema Resuelto"],
                    "Resuelto (1/0)": survey["Resuelto (1/0)"],
                    "Volveria a Usar": survey["Volveria a Usar"],
                    "Volveria (1/0)": survey["Volveria (1/0)"],
                    "NPS Categoria": survey["NPS Categoria"],
                    "Comentario": survey["Comentario"],
                    "Agente ID": row.get("agentId", ""),
                    "Conv ID": row.get("agentConversationId", ""),
                    "Close Reason": row.get("agentCloseReason", ""),
                }
                rows.append(processed)

    return rows


def write_excel(rows: list, output_path: Path):
    """Write processed rows to styled Excel file."""
    if not HAS_OPENPYXL:
        # Fallback: write CSV
        csv_path = output_path.with_suffix(".csv")
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
            if rows:
                writer = csv.DictWriter(f, fieldnames=rows[0].keys())
                writer.writeheader()
                writer.writerows(rows)
        print(f"openpyxl not installed. CSV saved to: {csv_path}")
        return csv_path

    wb = Workbook()
    ws = wb.active
    ws.title = "Encuestas NPS"

    if not rows:
        wb.save(output_path)
        return output_path

    headers = list(rows[0].keys())

    # Styles
    header_font = Font(name="Inter", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # Category fills
    promotor_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    pasivo_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    detractor_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data
    data_font = Font(name="Inter", size=10)
    nps_col_idx = headers.index("NPS Categoria") + 1

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            val = row_data[header]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

        # Color NPS category
        nps_cell = ws.cell(row=row_idx, column=nps_col_idx)
        if nps_cell.value == "Promotor":
            nps_cell.fill = promotor_fill
        elif nps_cell.value == "Pasivo":
            nps_cell.fill = pasivo_fill
        elif nps_cell.value == "Detractor":
            nps_cell.fill = detractor_fill

    # Auto-width columns
    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        for row_data in rows[:50]:
            val = str(row_data[header])
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 35)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    # === Summary sheet ===
    ws2 = wb.create_sheet("Resumen NPS")

    # NPS by month
    from collections import Counter
    monthly = {}
    for r in rows:
        m = r["Mes"]
        cat = r["NPS Categoria"]
        monthly.setdefault(m, Counter())
        monthly[m][cat] += 1

    ws2.cell(row=1, column=1, value="Mes").font = Font(bold=True, size=11)
    ws2.cell(row=1, column=2, value="Promotores").font = Font(bold=True, size=11)
    ws2.cell(row=1, column=3, value="Pasivos").font = Font(bold=True, size=11)
    ws2.cell(row=1, column=4, value="Detractores").font = Font(bold=True, size=11)
    ws2.cell(row=1, column=5, value="Total").font = Font(bold=True, size=11)
    ws2.cell(row=1, column=6, value="NPS Score").font = Font(bold=True, size=11)

    for i, (m, counts) in enumerate(sorted(monthly.items()), 2):
        total = sum(counts.values())
        prom = counts.get("Promotor", 0)
        pas = counts.get("Pasivo", 0)
        det = counts.get("Detractor", 0)
        nps = round((prom - det) / total * 100, 1) if total > 0 else 0

        ws2.cell(row=i, column=1, value=m)
        ws2.cell(row=i, column=2, value=prom)
        ws2.cell(row=i, column=3, value=pas)
        ws2.cell(row=i, column=4, value=det)
        ws2.cell(row=i, column=5, value=total)
        ws2.cell(row=i, column=6, value=nps)

    for col in range(1, 7):
        ws2.column_dimensions[get_column_letter(col)].width = 15

    wb.save(output_path)
    return output_path


def main():
    print("Processing NPS survey CSVs...")
    rows = process_all_csvs()
    print(f"Total survey responses: {len(rows)}")

    if not rows:
        print("No survey data found!")
        return

    # Print summary
    from collections import Counter
    cats = Counter(r["NPS Categoria"] for r in rows)
    total = len(rows)
    prom = cats.get("Promotor", 0)
    det = cats.get("Detractor", 0)
    nps = round((prom - det) / total * 100, 1) if total > 0 else 0

    print(f"\n{'='*50}")
    print(f"  Promotores:  {prom} ({prom/total*100:.1f}%)")
    print(f"  Pasivos:     {cats.get('Pasivo', 0)} ({cats.get('Pasivo', 0)/total*100:.1f}%)")
    print(f"  Detractores: {det} ({det/total*100:.1f}%)")
    print(f"  NPS Score:   {nps}")
    print(f"{'='*50}")

    # Print sample
    print("\nSample (first 3 rows):")
    for r in rows[:3]:
        print(f"  {r['Fecha']} | {r['Contacto'][:20]:20s} | {r['Entidad'][:20]:20s} | "
              f"Atencion={r['Score Atencion']} Asesor={r['Score Asesor']} | {r['NPS Categoria']}")

    # Print by entity
    entities = Counter(r["Entidad"] for r in rows)
    print(f"\nTop entidades:")
    for ent, count in entities.most_common(10):
        print(f"  {ent}: {count}")

    # Write output
    output = BASE_DIR / "encuestas_procesadas.xlsx"
    result = write_excel(rows, output)
    print(f"\nOutput: {result}")


if __name__ == "__main__":
    main()
